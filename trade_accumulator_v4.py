#!/usr/bin/env python3
"""
Trade Volume Accumulator  v4
============================
Parses commodity trade blotter screengrabs via Claude Vision and accumulates
volume + pricing data throughout a trading day into a structured Excel workbook.

Scope (v4)
----------
  - Outright trades
  - Same-product spreads (two or more legs, identical CC)
  - Same-timestamp / same-CC / different-qty rows → flagged as unrelated outrights
  - Cross-product situations (different CCs at same timestamp) → each CC treated
    as its own independent trade; no cross-spread logic applied

Usage
-----
    python trade_accumulator.py screenshot.png --output trade_tally.xlsx
    python trade_accumulator.py screenshot.png --dry-run   # preview only

Environment
-----------
    ANTHROPIC_API_KEY   or pass via --api-key
"""

import argparse
import base64
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
#  CLAUDE VISION — raw row extraction
# ═══════════════════════════════════════════════════════════════════════════

VISION_PROMPT = """
You are parsing a commodity trade blotter screenshot.

Return every visible data row as a JSON array — one object per row.
Do NOT group rows. Do NOT interpret spreads. Just return what you see.

Each object:
  "timestamp"   : string   e.g. "16:13:11 BST"
  "cc"          : string   e.g. "APC"
  "qty"         : integer
  "strip"       : string   e.g. "Dec26"
  "hub"         : string   e.g. "CIF ARA"
  "price"       : float
  "is_diff_row" : boolean  — true if this row looks like a spread differential
                             (small absolute value, same timestamp as adjacent
                             leg rows, sometimes negative). False otherwise.

Return ONLY a JSON array. No markdown, no explanation. Ignore header rows.
"""


def parse_image_with_claude(image_path: str, api_key: str) -> list[dict]:
    import urllib.request

    with open(image_path, "rb") as f:
        img_b64 = base64.standard_b64encode(f.read()).decode()

    ext  = Path(image_path).suffix.lower()
    mime = {".png": "image/png", ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg"}.get(ext, "image/png")

    payload = json.dumps({
        "model": "claude-opus-4-5-20251101",
        "max_tokens": 8192,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image",
                 "source": {"type": "base64", "media_type": mime, "data": img_b64}},
                {"type": "text", "text": VISION_PROMPT}
            ]
        }]
    }).encode()

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "x-api-key":         api_key,
            "anthropic-version": "2023-06-01",
            "content-type":      "application/json",
        },
        method="POST"
    )
    with urllib.request.urlopen(req) as resp:
        data = json.loads(resp.read())

    raw = "".join(b["text"] for b in data["content"] if b["type"] == "text")
    raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
    return json.loads(raw)


# ═══════════════════════════════════════════════════════════════════════════
#  TAPS / MOC DETECTION
# ═══════════════════════════════════════════════════════════════════════════

# Outrights for eligible CCs before this time at qualifying price → TAPS
TAPS_CUTOFF = "09:30:00"

# Exact prices that trigger TAPS classification, per CC group.
# Prices are rounded to 3 d.p. before comparison to absorb OCR float noise.
_SM_TAPS_PRICES = {-0.020, -0.015, -0.010, -0.005, 0.000, +0.005, +0.010, +0.015, +0.020}
_NJ_TAPS_PRICES = {-0.100, -0.050, 0.000, +0.050, +0.100}

TAPS_GROUPS = {
    "SMT": _SM_TAPS_PRICES,
    "SMU": _SM_TAPS_PRICES,
    "SMV": _SM_TAPS_PRICES,
    "SMS": _SM_TAPS_PRICES,
    "NJC": _NJ_TAPS_PRICES,
    "NJD": _NJ_TAPS_PRICES,
    "NJM": _NJ_TAPS_PRICES,
    "NJB": _NJ_TAPS_PRICES,
}
TAPS_CC = set(TAPS_GROUPS.keys())


def _classify_trade_type(trade: dict) -> str:
    """Return 'TAPS' if the trade meets MOC/TAPS criteria, else the original type."""
    if trade.get("trade_type") != "OUTRIGHT":
        return trade["trade_type"]
    cc = trade.get("cc", "")
    if cc not in TAPS_GROUPS:
        return "OUTRIGHT"
    time_part = trade.get("timestamp", "").split()[0]   # "HH:MM:SS"
    if time_part > TAPS_CUTOFF:
        return "OUTRIGHT"
    legs = trade.get("legs", [])
    if not legs:
        return "OUTRIGHT"
    price = round(float(legs[0].get("price", 999)), 3)
    if price in TAPS_GROUPS[cc]:
        return "TAPS"
    return "OUTRIGHT"


# ═══════════════════════════════════════════════════════════════════════════
#  STRIP SORT KEY — for sequential leg ordering
# ═══════════════════════════════════════════════════════════════════════════

_MONTH_IDX = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}


def _strip_sort_key(strip: str) -> tuple:
    """
    Return a comparable tuple so strips sort chronologically.
    Bal Month/ND → always first (front month).
    Cal YY       → (1, year, 99)   — full calendar year
    MonthYY-MonthYY → (1, start_year, start_month)  — strip range
    MonthYY      → (1, year, month_index)
    Q[1-4] YY   → (1, year, (q-1)*3)
    YY           → (1, year, 99)
    unknown      → (9, 99, 99)
    """
    s = strip.strip()
    if s.lower().startswith("bal"):
        return (0, 0, 0)
    # Cal YY — full calendar year
    m = re.match(r"^[Cc]al\s*(\d{2})$", s)
    if m:
        return (1, int(m.group(1)), 99)
    # Hyphenated strip range: Mar27-Jun27 — sort by start month
    m = re.match(r"^([A-Za-z]{3})(\d{2})-[A-Za-z]{3}\d{2}$", s)
    if m:
        mon = _MONTH_IDX.get(m.group(1).capitalize(), 99)
        return (1, int(m.group(2)), mon)
    # MonthYY
    m = re.match(r"^([A-Za-z]{3})(\d{2})$", s)
    if m:
        mon = _MONTH_IDX.get(m.group(1).capitalize(), 99)
        return (1, int(m.group(2)), mon)
    # Q[1-4] YY
    m = re.match(r"^Q([1-4])\s*(\d{2})?$", s)
    if m:
        return (1, int(m.group(2)) if m.group(2) else 99, (int(m.group(1)) - 1) * 3)
    # Bare year
    m = re.match(r"^(\d{2})$", s)
    if m:
        return (1, int(m.group(1)), 99)
    return (9, 99, 99)


# ── Contract lot sizes and display units ────────────────────────────────────────
# (lot_size, display_unit)  — display_unit is what the Vol Equiv column shows
CC_LOT: dict[str, tuple[int, str]] = {
    "AEB": (1_000, "bbl"),
    "AEN": (100,   "MT"),
    "AEO": (1_000, "MT"),
    "AOM": (100,   "MT"),
    "EOB": (1_000, "bbl"),  # traded MT, displayed bbl (×8.33 below)
    "EON": (1_000, "MT"),
    "EOO": (100,   "MT"),
    "GCZ": (100,   "MT"),
    "GDA": (100,   "MT"),
    "GDE": (1_000, "MT"),
    "GDK": (1_000, "bbl"),
    "GDO": (1_000, "MT"),
    "GDQ": (100,   "MT"),
    "NAM": (100,   "MT"),
    "NBB": (1_000, "bbl"),
    "NEC": (1_000, "MT"),
    "NEH": (1_000, "MT"),
    "NJC": (1_000, "MT"),
    "NJD": (1_000, "MT"),
    "NOB": (1_000, "bbl"),  # traded MT, displayed bbl (×8.90 below)
    "SMD": (1_000, "bbl"),
    "SMF": (1_000, "bbl"),
    "SMJ": (1_000, "bbl"),
    "SMT": (1_000, "bbl"),
    "SMV": (100,   "bbl"),
    "STB": (1_000, "bbl"),
    "UCB": (1_000, "MT"),
    # ── Freight (1 lot = 1,000 MT = 1 KT; vol equiv displayed in KT) ─────
    "JFF": (1, "KT"),
    "TDC": (1, "KT"),
    "TDA": (1, "KT"),
    "TDL": (1, "KT"),
    "WAT": (1, "KT"),
    "WDF": (1, "KT"),
    "WFA": (1, "KT"),
    "WHK": (1, "KT"),
    "WMJ": (1, "KT"),
    "WNS": (1, "KT"),
    "WNX": (1, "KT"),
    "WSN": (1, "KT"),
}

# Contracts traded in MT but quoted/displayed in bbl: multiply MT vol by this factor
CC_MT_TO_BBL: dict[str, float] = {
    "EOB": 8.33,
    "NOB": 8.90,
}

# Balmo CCs amalgamated into parent CC in the tally
CC_ALIAS: dict[str, str] = {
    "TDA": "TDC",   # TD20 Balmo → TD20
    "WNX": "WMJ",   # TC5 Balmo  → TC5
}

# ── Strip-range month counter ───────────────────────────────────────────────────
_RANGE_RE = re.compile(r"^([A-Za-z]{3})(\d{2})-([A-Za-z]{3})(\d{2})$")


def _strip_multiplier(strip: str) -> int:
    """
    Return the volume-equivalent multiplier for a strip designation.
    Cal = 12, Quarter = 3, MonthYY-MonthYY = number of months, else 1.
    """
    s = strip.strip()
    if re.match(r"^[Cc]al", s):
        return 12
    if re.match(r"^Q[1-4]", s, re.IGNORECASE):
        return 3
    m = _RANGE_RE.match(s)
    if m:
        mon0 = _MONTH_IDX.get(m.group(1).capitalize(), 0)
        yr0  = int(m.group(2))
        mon1 = _MONTH_IDX.get(m.group(3).capitalize(), 0)
        yr1  = int(m.group(4))
        return max(1, (yr1 - yr0) * 12 + (mon1 - mon0) + 1)
    return 1


def _vol_equiv(cc: str, qty: float, strip: str, tt: str) -> tuple[float, str]:
    """
    Return (vol_equiv, unit) for a trade.
    vol_equiv = qty × lot_size × strip_mult [× MT→bbl factor if applicable]
                × 2 for BUTTERFLY / CONDOR
    """
    lot_size, unit = CC_LOT.get(cc, (1, "lot"))
    sm = _strip_multiplier(strip)
    if tt in ("BUTTERFLY", "CONDOR"):
        sm *= 2
    veq = qty * lot_size * sm
    if cc in CC_MT_TO_BBL:
        veq  = veq * CC_MT_TO_BBL[cc]
        unit = "bbl"
    return round(veq, 2), unit


def _volume_multiplier(trade: dict) -> int:
    """
    Return the volume-equivalent multiplier for a trade.
      OUTRIGHT / TAPS / SPREAD / INTERPRODUCT → strip multiplier (of first leg)
      BUTTERFLY / CONDOR                       → strip multiplier × 2
    """
    tt   = trade.get("trade_type", "")
    legs = trade.get("legs", [])
    strip = legs[0].get("strip", "") if legs else ""
    sm = _strip_multiplier(strip)
    if tt in ("BUTTERFLY", "CONDOR"):
        return sm * 2
    return sm


def _next_month_strip() -> str:
    """
    Return the MonthYY token for the calendar month after the current one.
    June 2026 → "Jul26", December 2026 → "Jan27", etc.
    Used to synthesise the missing leg of a Bal Month spread.
    """
    from datetime import date
    today = date.today()
    if today.month == 12:
        nm_month, nm_year = 1, today.year + 1
    else:
        nm_month, nm_year = today.month + 1, today.year
    month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                   "Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{month_names[nm_month - 1]}{str(nm_year)[2:]}"


def _try_synthesise_spread(leg_row: dict, diff_row: dict,
                           ts: str, cc: str) -> dict | None:
    """
    When there is exactly one price leg and one diff row, synthesise the
    missing leg and return a SPREAD trade dict, or None if the strips don't
    align.

    Diff strip must be "A/B".  If we hold leg A:  missing B = A − diff.
                                If we hold leg B:  missing A = B + diff.

    Special case: Bal Month spreads in the new blotter format show the diff
    row with strip "Bal Month" (or "Bal Month-ND") and no "/" notation.
    The second leg is ALWAYS the next calendar month.  We expand the strip
    here before the standard matching logic.
    """
    diff_strip = diff_row["strip"].strip()

    # Expand bare Bal Month diff strips to "Bal Month/NextMonth".
    if "/" not in diff_strip and diff_strip.lower().startswith("bal"):
        diff_strip = f"{diff_strip}/{_next_month_strip()}"

    parts = diff_strip.split("/")
    if len(parts) != 2:
        return None
    leg_strip = leg_row["strip"].strip()
    d_a, d_b  = parts[0].strip(), parts[1].strip()
    if d_a == leg_strip:
        missing_strip = d_b
        missing_price = round(leg_row["price"] - diff_row["price"], 6)
    elif d_b == leg_strip:
        missing_strip = d_a
        missing_price = round(leg_row["price"] + diff_row["price"], 6)
    else:
        return None
    legs = sorted(
        [{"strip": leg_strip,    "price": leg_row["price"]},
         {"strip": missing_strip, "price": missing_price}],
        key=lambda l: _strip_sort_key(l["strip"]),
    )
    return {
        "timestamp":    ts,
        "trade_type":   "SPREAD",
        "notes":        "leg synthesised",
        "cc":           cc,
        "qty":          leg_row["qty"],
        "hub":          leg_row.get("hub", ""),
        "spread_price": diff_row["price"],
        "legs":         legs,
    }


def _assign_balmo_diff_cc(ts_rows: list[dict]) -> list[dict]:
    """
    Pre-processing pass for a single timestamp's rows.

    Bal Month spread diff rows (strip like "Bal Month/Jul26") often have a
    blank CC because the blotter omits it on diff lines.  If we can find a
    non-diff leg at the same timestamp whose strip starts with "Bal" and
    whose CC is non-blank, we copy that CC onto the blank-CC diff row so
    that both rows land in the same cc_map bucket.

    Also handles "Bal Month-ND" and other "Bal …" variants.
    """
    # Build a map: bal-strip-prefix → CC  from known leg rows
    # e.g. "Bal Month" → "NJD",  "Bal Month-ND" → "SMU"
    bal_leg_cc: dict[str, str] = {}
    for r in ts_rows:
        if (not r.get("is_diff_row") and not r.get("cancelled")
                and r.get("cc") and r["strip"].lower().startswith("bal")):
            bal_leg_cc[r["strip"].strip()] = r["cc"]

    if not bal_leg_cc:
        return ts_rows

    result = []
    for r in ts_rows:
        if r.get("is_diff_row") and r["strip"].lower().startswith("bal"):
            # Always inherit CC from the matching Bal leg, overriding any
            # hub-derived CC (hub lookup can return a sibling CC like NJC
            # when the leg is NJD — both share the same hub name).
            first_part = r["strip"].split("/")[0].strip()
            cc = bal_leg_cc.get(first_part) or (
                next(iter(bal_leg_cc.values())) if len(bal_leg_cc) == 1 else "")
            if cc:
                r = dict(r)   # shallow copy — don't mutate the original
                r["cc"] = cc
        result.append(r)
    return result


# ═══════════════════════════════════════════════════════════════════════════
#  TRADE GROUPING
# ═══════════════════════════════════════════════════════════════════════════

def group_rows_into_trades(raw_rows: list[dict]) -> list[dict]:
    """
    Group raw blotter rows into trades.

    Primary key: (Ex.Time, Sub.Time).  Rows sharing both timestamps are
    candidate legs of the same multi-leg trade.  Sub.Time="" for old-
    format rows without a submission timestamp.

    Within each (Ex.Time, Sub.Time) group:

      Cross-CC pass first:
        • 2 active legs, different CCs, same qty, no diff → INTERPRODUCT_SPREAD

      Per-CC:
        1. 1 leg + diff row        → SPREAD (missing leg synthesised)
        2. 1 leg, no diff          → OUTRIGHT / TAPS
        3. 4 legs, all equal qty   → CONDOR
        4. 3 legs, outer N mid 2N  → BUTTERFLY
        5. 2+ legs, equal qty      → SPREAD
        6. 2 legs, unequal qty     → SPREAD (implied diff)
        7. 3+ legs, unresolvable   → flagged OUTRIGHT per leg

    Legs are always sorted into chronological strip order.
    """
    from itertools import groupby as _groupby

    rows = sorted(raw_rows,
                  key=lambda r: (r["timestamp"], r.get("sub_time", "")))
    trades: list[dict] = []

    for grp_key, grp in _groupby(
            rows, key=lambda r: (r["timestamp"], r.get("sub_time", ""))):
        ts      = grp_key[0]
        ts_rows = list(grp)

        # Assign blank-CC Bal Month diff rows to the matching leg's CC
        ts_rows = _assign_balmo_diff_cc(ts_rows)

        # One CANCELLED trade per cancelled row
        for cr in ts_rows:
            if cr.get("cancelled"):
                trades.append({
                    "timestamp":    ts,
                    "trade_type":   "CANCELLED",
                    "notes":        "Cancelled trade — excluded from tally",
                    "cc":           cr.get("cc", ""),
                    "qty":          cr["qty"],
                    "hub":          cr.get("hub", ""),
                    "spread_price": None,
                    "legs": [{"strip": cr["strip"], "price": cr["price"]}],
                })

        active    = [r for r in ts_rows if not r.get("cancelled")]
        diffs_all = [r for r in active if r.get("is_diff_row")]
        legs_all  = [r for r in active if not r.get("is_diff_row")]

        # ── Handle standalone diff/fly rows (no separate leg rows present) ──
        # New blotter format emits a single row for spreads and butterflies.
        # Process these before the cross-CC / per-CC logic that expects legs.
        standalone_diffs = []
        for dr in diffs_all:
            parts = dr["strip"].split("/")
            strat = dr.get("strategy", "")
            if strat == "fly" and len(parts) == 3:
                # Butterfly: single fly row with 3-strip notation.
                leg_strips = sorted(parts, key=_strip_sort_key)
                trades.append({
                    "timestamp":    ts,
                    "trade_type":   "BUTTERFLY",
                    "notes":        "fly row",
                    "cc":           dr["cc"],
                    "qty":          dr["qty"],
                    "hub":          dr.get("hub", ""),
                    "spread_price": dr["price"],
                    "legs":         [{"strip": s} for s in leg_strips],
                })
            elif len(parts) == 2:
                # Spread diff row — check if matching outright legs exist;
                # if not, emit as a standalone spread (price is the differential).
                matched_leg = next(
                    (lr for lr in legs_all
                     if lr["cc"] == dr["cc"]
                     and lr["strip"] in parts),
                    None,
                )
                if matched_leg is None:
                    # No leg rows — emit the spread directly from the diff row.
                    leg_strips = sorted(parts, key=_strip_sort_key)
                    trades.append({
                        "timestamp":    ts,
                        "trade_type":   "SPREAD",
                        "notes":        "spread row",
                        "cc":           dr["cc"],
                        "qty":          dr["qty"],
                        "hub":          dr.get("hub", ""),
                        "spread_price": dr["price"],
                        "legs":         [{"strip": s} for s in leg_strips],
                    })
                else:
                    standalone_diffs.append(dr)   # has matching legs — pair normally
            else:
                standalone_diffs.append(dr)

        diffs_all = standalone_diffs

        if not legs_all:
            continue

        # ── Cross-CC: interproduct spread ─────────────────────────────────
        leg_ccs = list({r["cc"] for r in legs_all if r["cc"]})
        if (len(legs_all) == 2 and len(leg_ccs) == 2
                and legs_all[0]["qty"] == legs_all[1]["qty"]
                and not diffs_all):
            ls = sorted(legs_all, key=lambda r: r["cc"])
            sp = round(ls[0]["price"] - ls[1]["price"], 6)
            trades.append({
                "timestamp":    ts,
                "trade_type":   "INTERPRODUCT_SPREAD",
                "notes":        "implied diff",
                "cc":           f"{ls[0]['cc']}/{ls[1]['cc']}",
                "qty":          ls[0]["qty"],
                "hub":          ls[0]["hub"],
                "spread_price": sp,
                "legs": [{"strip": r["strip"], "price": r["price"], "cc": r["cc"]}
                          for r in ls],
            })
            continue

        # ── Per-CC grouping ───────────────────────────────────────────────
        cc_map: dict[str, list] = {}
        for row in active:
            cc_map.setdefault(row["cc"], []).append(row)

        for cc, cc_rows in cc_map.items():
            diff_rows = [r for r in cc_rows if r.get("is_diff_row")]
            leg_rows  = [r for r in cc_rows if not r.get("is_diff_row")]

            if not leg_rows:
                continue

            # ── Single leg ────────────────────────────────────────────────
            if len(leg_rows) == 1:
                lr    = leg_rows[0]
                synth = (_try_synthesise_spread(lr, diff_rows[0], ts, cc)
                         if diff_rows else None)
                if synth:
                    trades.append(synth)
                else:
                    t = {
                        "timestamp":    ts,
                        "trade_type":   "OUTRIGHT",
                        "notes":        "",
                        "cc":           cc,
                        "qty":          lr["qty"],
                        "hub":          lr.get("hub", ""),
                        "spread_price": None,
                        "legs": [{"strip": lr["strip"], "price": lr["price"]}],
                    }
                    t["trade_type"] = _classify_trade_type(t)
                    if t["trade_type"] == "TAPS":
                        t["notes"] = "TAPS/MOC"
                    trades.append(t)

            # ── Multiple legs ─────────────────────────────────────────────
            else:
                ls   = sorted(leg_rows, key=lambda r: _strip_sort_key(r["strip"]))
                qtys = [r["qty"] for r in ls]

                # Condor: 4 equal-qty legs
                if len(ls) == 4 and len(set(qtys)) == 1:
                    p = [r["price"] for r in ls]
                    condor_px = round((p[0] - p[1]) - (p[2] - p[3]), 6)
                    trades.append({
                        "timestamp":    ts,
                        "trade_type":   "CONDOR",
                        "notes":        "condor",
                        "cc":           cc,
                        "qty":          qtys[0],
                        "hub":          ls[0]["hub"],
                        "spread_price": condor_px,
                        "legs": [{"strip": r["strip"], "price": r["price"]} for r in ls],
                    })

                # Butterfly: 3 legs, outer equal, middle = 2× outer
                elif (len(ls) == 3
                      and qtys[0] == qtys[2]
                      and qtys[1] == 2 * qtys[0]):
                    p   = [r["price"] for r in ls]
                    fly = round((p[0] - p[1]) - (p[1] - p[2]), 6)
                    trades.append({
                        "timestamp":    ts,
                        "trade_type":   "BUTTERFLY",
                        "notes":        "butterfly",
                        "cc":           cc,
                        "qty":          qtys[0],
                        "hub":          ls[0]["hub"],
                        "spread_price": fly,
                        "legs": [{"strip": r["strip"], "price": r["price"]} for r in ls],
                    })

                # Same qty → SPREAD
                elif len(set(qtys)) == 1:
                    sp   = (diff_rows[0]["price"] if diff_rows
                            else round(ls[0]["price"] - ls[1]["price"], 6))
                    note = "" if diff_rows else "implied diff"
                    trades.append({
                        "timestamp":    ts,
                        "trade_type":   "SPREAD",
                        "notes":        note,
                        "cc":           cc,
                        "qty":          qtys[0],
                        "hub":          ls[0]["hub"],
                        "spread_price": sp,
                        "legs": [{"strip": r["strip"], "price": r["price"]} for r in ls],
                    })

                # 2 legs, unequal qty → SPREAD with implied diff
                elif len(ls) == 2:
                    sp = round(ls[0]["price"] - ls[1]["price"], 6)
                    trades.append({
                        "timestamp":    ts,
                        "trade_type":   "SPREAD",
                        "notes":        "implied diff (unequal qty)",
                        "cc":           cc,
                        "qty":          ls[0]["qty"],
                        "hub":          ls[0]["hub"],
                        "spread_price": sp,
                        "legs": [{"strip": r["strip"], "price": r["price"]} for r in ls],
                    })

                # 3+ legs, unresolvable mixed qty → flag each
                else:
                    for lr in leg_rows:
                        trades.append({
                            "timestamp":    ts,
                            "trade_type":   "OUTRIGHT",
                            "notes":        "⚠ FLAG: same-timestamp same-CC different qty — verify",
                            "cc":           cc,
                            "qty":          lr["qty"],
                            "hub":          lr.get("hub", ""),
                            "spread_price": None,
                            "legs": [{"strip": lr["strip"], "price": lr["price"]}],
                        })

    return trades


# ═══════════════════════════════════════════════════════════════════════════
#  STYLING
# ═══════════════════════════════════════════════════════════════════════════

F_HDR       = PatternFill("solid", fgColor="1F3864")  # navy        — headers
F_OUT       = PatternFill("solid", fgColor="FFFFFF")  # white       — outrights
F_SPREAD    = PatternFill("solid", fgColor="FFF2CC")  # amber       — spreads
F_FLAG      = PatternFill("solid", fgColor="FCE4D6")  # salmon      — flagged
F_CC_BAND   = PatternFill("solid", fgColor="1F3864")  # navy        — CC banner
F_GRAND     = PatternFill("solid", fgColor="2E75B6")  # mid-blue    — Grand Total row
F_BLK_HDR   = PatternFill("solid", fgColor="D6E4F0")  # steel blue  — block header
F_TRADE_ALT = PatternFill("solid", fgColor="EBF3FB")  # pale blue   — alt trade rows
F_SUMMARY   = PatternFill("solid", fgColor="E2EFDA")  # pale green  — VWAP summary

FN_HDR   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FN_BODY  = Font(name="Arial", size=10)
FN_BOLD  = Font(name="Arial", bold=True, size=10)
FN_FLAG  = Font(name="Arial", size=10, color="C00000")
FN_TITLE = Font(name="Arial", bold=True, size=13, color="1F3864")

_T = Side(style="thin",   color="BDD7EE")
_M = Side(style="medium", color="2E75B6")
BORD  = Border(left=_T, right=_T, top=_T, bottom=_T)


def _hdr(ws, row_num, labels, fill=None):
    for ci, label in enumerate(labels, 1):
        c = ws.cell(row=row_num, column=ci, value=label)
        c.font      = FN_HDR
        c.fill      = fill or F_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORD
    ws.row_dimensions[row_num].height = 30


def _c(ws, r, c_idx, val, fill=None, halign="left",
       bold=False, flag=False, font=None):
    cell = ws.cell(row=r, column=c_idx, value=val)
    cell.fill      = fill or F_OUT
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border    = BORD
    if flag:
        cell.font = FN_FLAG
    elif bold:
        cell.font = FN_BOLD
    else:
        cell.font = font or FN_BODY
    return cell


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════

SH_LOG     = "Trade Log"
SH_TALLY   = "Volume Tally"
SH_IMPORT  = "Import Log"
SH_SUMMARY = "Day Summary"

# Freight CC categorisation for Day Summary
FREIGHT_CLEAN_CC = {"WMJ", "WSN", "WHK", "JFF", "WNX", "WNS"}
FREIGHT_DIRTY_CC = {"TDC", "TDL", "WDF", "TDA"}
FREIGHT_LPG_CC   = {"WAT", "WFA"}

# Trade Log columns
# A  Timestamp | B  Trade Type | C  Notes/Flags | D  CC | E  Qty | F  Hub
# G  Spread Price
# H  Leg 1 Strip | I  Leg 1 Price
# J  Leg 2 Strip | K  Leg 2 Price
# L  Leg 3 Strip | M  Leg 3 Price
# N  Source File

LOG_COLS = [
    "Timestamp", "Trade Type", "Notes / Flags", "CC", "Qty", "Hub",
    "Spread Price",
    "Leg 1 Strip", "Leg 1 Price",
    "Leg 2 Strip", "Leg 2 Price",
    "Leg 3 Strip", "Leg 3 Price",
    "Source File",
]
LOG_WIDTHS = {
    "A": 16, "B": 12, "C": 40, "D": 8,  "E": 7,  "F": 16,
    "G": 14,
    "H": 14, "I": 12,
    "J": 14, "K": 12,
    "L": 14, "M": 12,
    "N": 28,
}

# Volume Tally columns
# A  Block label (CC × Strip or CC × Spread pair)
# B  Timestamp (or summary label)
# C  Qty
# D  Price  (per-trade price — never averaged)
# E  Cumul. Volume  (running)
# F  VWAP           (running)

TALLY_COLS   = ["CC × Strip / Spread", "Timestamp", "Qty", "Vol Equiv",
                 "Price", "High", "Low", "Cumul. Vol", "Cumul. Vol Equiv", "VWAP (running)"]
TALLY_WIDTHS = {"A": 36, "B": 18, "C": 10, "D": 12,
                 "E": 14, "F": 12, "G": 12, "H": 14, "I": 16, "J": 16}
TALLY_NUM_COLS = 10  # total column count


# ═══════════════════════════════════════════════════════════════════════════
#  BUILD FRESH WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════

def build_fresh_workbook(path: str):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(SH_LOG)
    ws.freeze_panes = "A2"
    _hdr(ws, 1, LOG_COLS)
    for col, w in LOG_WIDTHS.items():
        ws.column_dimensions[col].width = w

    wt = wb.create_sheet(SH_TALLY)
    wt.freeze_panes = "A3"
    wt["A1"] = "Volume Tally  —  per-trade prices · vol equiv · cumulative volume · VWAP"
    wt["A1"].font      = FN_TITLE
    wt["A1"].alignment = Alignment(horizontal="left", vertical="center")
    wt.row_dimensions[1].height = 28
    _hdr(wt, 2, TALLY_COLS)
    for col, w in TALLY_WIDTHS.items():
        wt.column_dimensions[col].width = w

    ws2 = wb.create_sheet(SH_SUMMARY)
    ws2.freeze_panes = "A3"
    ws2["A1"] = "Day Summary  —  Freight volumes by category"
    ws2["A1"].font      = FN_TITLE
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 28

    wi = wb.create_sheet(SH_IMPORT)
    _hdr(wi, 1, ["Import Time", "Source File", "Raw Rows",
                  "Trades", "New Added", "Skipped"])
    for col, w in {"A": 22, "B": 32, "C": 12,
                    "D": 10, "E": 12, "F": 10}.items():
        wi.column_dimensions[col].width = w

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════
#  DEDUPLICATION KEY
# ═══════════════════════════════════════════════════════════════════════════

def load_existing_keys(path: str) -> set:
    """
    (timestamp, trade_type, cc, leg1_strip) — four fields to safely
    distinguish two legitimate trades of the same type/CC at the same second.
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[SH_LOG]
        keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[3] and row[7]:
                keys.add((str(row[0]).strip(), str(row[1]).strip(),
                          str(row[3]).strip(), str(row[7]).strip()))
        wb.close()
        return keys
    except Exception:
        return set()


# ═══════════════════════════════════════════════════════════════════════════
#  WRITE ONE TRADE LOG ROW
# ═══════════════════════════════════════════════════════════════════════════

def _write_log_row(ws, t: dict, source_file: str):
    legs = t.get("legs", [])
    tt   = t["trade_type"]
    note = t.get("notes", "")

    fill = (F_FLAG   if "⚠" in note else
            F_SPREAD if tt == "SPREAD" else
            F_OUT)

    r = ws.max_row + 1
    ws.row_dimensions[r].height = 17

    def leg(i, field):
        return legs[i].get(field) if i < len(legs) else None

    vals = [
        t["timestamp"], tt, note, t["cc"], t["qty"], t.get("hub", ""),
        t.get("spread_price"),
        leg(0, "strip"), leg(0, "price"),
        leg(1, "strip"), leg(1, "price"),
        leg(2, "strip"), leg(2, "price"),
        source_file,
    ]

    RIGHT = {5, 7, 9, 11, 13}  # Qty + price columns (1-indexed)
    for ci, val in enumerate(vals, 1):
        _c(ws, r, ci, val,
           fill=fill,
           halign="right" if ci in RIGHT else "left",
           flag=(ci == 3 and "⚠" in str(val or "")))


# ═══════════════════════════════════════════════════════════════════════════
#  REBUILD VOLUME TALLY
# ═══════════════════════════════════════════════════════════════════════════

def _rebuild_tally(wb):
    ws = wb[SH_LOG]
    wt = wb[SH_TALLY]

    # ── Read all trades from Trade Log ─────────────────────────────────────
    all_trades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        ts    = str(row[0] or "")
        tt    = str(row[1] or "")
        cc    = str(row[3] or "")
        qty   = row[4] or 0
        hub   = str(row[5] or "")
        sp    = row[6]
        legs  = []
        for i in range(3):
            base = 7 + i * 2
            s, p = row[base], row[base + 1]
            if s:
                legs.append({"strip": s, "price": p})
        all_trades.append({
            "ts": ts, "tt": tt, "cc": cc, "hub": hub,
            "qty": qty or 0, "sp": sp, "legs": legs,
        })

    if not all_trades:
        return

    # ── Build buckets ──────────────────────────────────────────────────────
    from collections import defaultdict

    # key = (cc, strip_label, kind_str)
    # value = list of {"ts", "qty", "vol_equiv", "price"}
    buckets:    dict[tuple, list] = defaultdict(list)
    bucket_hub: dict[tuple, str]  = {}   # cc → hub name

    for t in all_trades:
        cc, qty, legs, tt, hub = (
            t["cc"], t["qty"], t["legs"], t["tt"], t["hub"])

        if tt == "CANCELLED" or not legs:
            continue

        # Amalgamate balmo CCs into their parent (e.g. TDA → TDC, WNX → WMJ)
        cc = CC_ALIAS.get(cc, cc)

        strip0 = legs[0].get("strip", "")
        veq, unit = _vol_equiv(cc, qty, strip0, tt)

        if tt in ("OUTRIGHT", "TAPS"):
            l   = legs[0]
            key = (cc, l["strip"], tt)
            buckets[key].append(
                {"ts": t["ts"], "qty": qty, "vol_equiv": veq,
                 "price": l["price"]})

        elif tt in ("SPREAD", "BUTTERFLY", "CONDOR", "INTERPRODUCT_SPREAD"):
            if t["sp"] is None:
                continue
            diff_label = " / ".join(
                l["strip"] for l in sorted(legs,
                                           key=lambda l: _strip_sort_key(l["strip"]))
            )
            key = (cc, diff_label, tt)
            buckets[key].append(
                {"ts": t["ts"], "qty": qty, "vol_equiv": veq,
                 "price": t["sp"]})

        if cc not in bucket_hub and hub:
            bucket_hub[cc] = hub

    # Sort each bucket chronologically
    for k in buckets:
        buckets[k].sort(key=lambda x: x["ts"])

    # ── Sort keys: CC → kind → strip label ────────────────────────────────
    KIND_ORDER = {"OUTRIGHT": 0, "TAPS": 1, "SPREAD": 2, "BUTTERFLY": 3,
                  "CONDOR": 4, "INTERPRODUCT_SPREAD": 5}
    sorted_keys = sorted(
        buckets.keys(),
        key=lambda k: (k[0], KIND_ORDER.get(k[2], 9), k[1])
    )

    # ── Clear old tally rows (row 3+) ─────────────────────────────────────
    for rw in wt.iter_rows(min_row=3):
        for c in rw:
            c.value  = None
            c.fill   = F_OUT
            c.border = Border()
            c.font   = FN_BODY

    NC = TALLY_NUM_COLS  # 10

    def _tally_row(row_num, vals, aligns, fill, bold=False, height=16):
        for ci, (val, ha) in enumerate(zip(vals, aligns), 1):
            _c(wt, row_num, ci, val, fill=fill, halign=ha, bold=bold)
        wt.row_dimensions[row_num].height = height

    def _banner_row(row_num, text, fill, font, height=22):
        for ci in range(1, NC + 1):
            c = wt.cell(row=row_num, column=ci,
                        value=(text if ci == 1 else None))
            c.fill      = fill
            c.font      = font
            c.border    = BORD
            c.alignment = Alignment(horizontal="left", vertical="center")
        wt.row_dimensions[row_num].height = height

    # ── Write blocks ──────────────────────────────────────────────────────
    r       = 3
    prev_cc = None

    # Accumulate per-CC strategy subtotals: {cc: {kind: {"vol": N, "vol_equiv": N}}}
    cc_strategy_vol: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {"vol": 0, "vol_equiv": 0}))

    RIGHT = {"right"}  # convenience

    for key in sorted_keys:
        cc_key, strip_label, kind = key
        recs = buckets[key]
        trade_fill = (F_SPREAD if kind in ("SPREAD", "BUTTERFLY", "CONDOR",
                                            "INTERPRODUCT_SPREAD")
                      else F_FLAG if kind == "TAPS" else F_OUT)

        # ── CC banner ──────────────────────────────────────────────────────
        if cc_key != prev_cc:
            if prev_cc is not None:
                # Strategy subtotals for previous CC
                _write_strategy_subtotals(wt, r, prev_cc,
                                          cc_strategy_vol[prev_cc], NC)
                r += len(cc_strategy_vol[prev_cc]) + 2  # subtotals + blank spacer

            hub_name = bucket_hub.get(cc_key, "")
            is_freight = cc_key in _ALL_FREIGHT_CC
            banner = f"{cc_key}  —  {hub_name}" if (hub_name and is_freight) else cc_key
            _banner_row(r, banner, F_CC_BAND, FN_HDR, height=22)
            r += 1
            prev_cc = cc_key

        # ── Block sub-header ───────────────────────────────────────────────
        KIND_LABELS = {
            "OUTRIGHT": "Outright", "TAPS": "TAPS / MOC",
            "SPREAD": "Spread", "BUTTERFLY": "Butterfly",
            "CONDOR": "Condor", "INTERPRODUCT_SPREAD": "Inter-product Spread",
        }
        block_title = f"{strip_label}  [{KIND_LABELS.get(kind, kind)}]"
        _banner_row(r, block_title, F_BLK_HDR, FN_BOLD, height=18)
        r += 1

        # ── Individual trade rows ──────────────────────────────────────────
        cumvol = cumveq = vwap_num = 0.0
        hi = lo = None

        for i, rec in enumerate(recs):
            qty_val   = rec["qty"]      or 0
            veq_val   = rec["vol_equiv"] or 0
            price_val = rec["price"]    or 0.0

            cumvol   += qty_val
            cumveq   += veq_val
            vwap_num += qty_val * price_val

            if price_val:
                hi = price_val if hi is None else max(hi, price_val)
                lo = price_val if lo is None else min(lo, price_val)

            running_vwap = round(vwap_num / cumvol, 6) if cumvol else None
            row_fill = F_TRADE_ALT if i % 2 == 0 else trade_fill

            # A=blank, B=ts, C=qty, D=vol_equiv, E=price, F=hi, G=lo,
            # H=cumvol, I=cumveq, J=vwap
            vals   = [None, rec["ts"], qty_val, veq_val, rec["price"],
                      None, None, cumvol, cumveq, running_vwap]
            aligns = ["left","left","right","right","right",
                      "right","right","right","right","right"]
            _tally_row(r, vals, aligns, row_fill, height=16)
            r += 1

            # Accumulate for strategy subtotals
            cc_strategy_vol[cc_key][kind]["vol"]      += qty_val
            cc_strategy_vol[cc_key][kind]["vol_equiv"] += veq_val

        # ── Summary / VWAP row for this block ─────────────────────────────
        trade_count  = len(recs)
        final_vwap   = round(vwap_num / cumvol, 6) if cumvol else None
        count_label  = f"► {trade_count} trade{'s' if trade_count != 1 else ''}  |  Cumul. Vol / VWAP"
        sum_vals     = [count_label, "", cumvol, cumveq, "",
                        hi, lo, cumvol, cumveq, final_vwap]
        sum_aligns   = ["left","left","right","right","right",
                        "right","right","right","right","right"]
        _tally_row(r, sum_vals, sum_aligns, F_SUMMARY, bold=True, height=18)
        r += 1

    # ── Final CC strategy subtotals ────────────────────────────────────────
    if prev_cc is not None:
        _write_strategy_subtotals(wt, r, prev_cc,
                                  cc_strategy_vol[prev_cc], NC)
        r += len(cc_strategy_vol[prev_cc]) + 1


def _write_strategy_subtotals(wt, row_start, cc_key, strategy_data, nc):
    """Write a compact strategy-volume breakdown block below a CC section."""
    KIND_LABELS = {
        "OUTRIGHT": "Outright", "TAPS": "TAPS / MOC",
        "SPREAD": "Spread", "BUTTERFLY": "Butterfly",
        "CONDOR": "Condor", "INTERPRODUCT_SPREAD": "Inter-product Spread",
    }
    r = row_start
    # Section header
    for ci in range(1, nc + 1):
        c = wt.cell(row=r, column=ci,
                    value=(f"{cc_key}  — Strategy Summary" if ci == 1 else None))
        c.fill      = F_BLK_HDR
        c.font      = FN_BOLD
        c.border    = BORD
        c.alignment = Alignment(horizontal="left", vertical="center")
    wt.row_dimensions[r].height = 16
    r += 1

    for kind, data in strategy_data.items():
        label = KIND_LABELS.get(kind, kind)
        vals  = [f"  {label}", "", data["vol"], data["vol_equiv"],
                 "", "", "", "", "", ""]
        for ci, val in enumerate(vals, 1):
            _c(wt, r, ci, val,
               fill=F_SUMMARY,
               halign="right" if ci in (3, 4) else "left",
               bold=False)
        wt.row_dimensions[r].height = 15
        r += 1


# ═══════════════════════════════════════════════════════════════════════════
#  DAY SUMMARY (Freight)
# ═══════════════════════════════════════════════════════════════════════════

_ALL_FREIGHT_CC = FREIGHT_CLEAN_CC | FREIGHT_DIRTY_CC | FREIGHT_LPG_CC

_SUMMARY_CATEGORIES = [
    ("Clean Tanker", FREIGHT_CLEAN_CC),
    ("Dirty Tanker", FREIGHT_DIRTY_CC),
    ("LPG",          FREIGHT_LPG_CC),
]

# A=CC, B=Unit, C=Outright Vol, D=Outright Vol Equiv, E=Spread Vol, F=Spread Vol Equiv,
# G=Total Vol, H=Total Vol Equiv, I=# Trades
_SUMMARY_HDR_COLS = [
    "CC", "Unit", "Outright Vol", "Outright Vol Eq",
    "Spread Vol", "Spread Vol Eq",
    "Total Vol", "Total Vol Eq", "# Trades",
]
_SUMMARY_HDR_WIDTHS = {
    "A": 14, "B": 8,  "C": 14, "D": 16,
    "E": 12, "F": 14, "G": 12, "H": 14, "I": 10,
}

_OUTRIGHT_TYPES = {"OUTRIGHT", "TAPS"}
_SPREAD_TYPES   = {"SPREAD", "BUTTERFLY", "CONDOR", "INTERPRODUCT_SPREAD"}


def _build_day_summary(wb):
    """Rebuild the Day Summary sheet: per-CC outright vs spread totals, all products."""
    ws  = wb[SH_LOG]
    ws2 = wb[SH_SUMMARY]

    from collections import defaultdict

    def _blank_cc():
        return {"out_vol": 0.0, "out_veq": 0.0,
                "spr_vol": 0.0, "spr_veq": 0.0,
                "trades":  0,   "unit": ""}

    cc_data: dict[str, dict] = defaultdict(_blank_cc)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        tt  = str(row[1] or "")
        cc  = str(row[3] or "")
        qty = row[4] or 0
        legs = []
        for i in range(3):
            base = 7 + i * 2
            s, p = row[base], row[base + 1]
            if s:
                legs.append({"strip": s, "price": p})

        if tt == "CANCELLED" or not legs:
            continue

        # Amalgamate balmo CCs into their parent
        cc = CC_ALIAS.get(cc, cc)

        strip0 = legs[0].get("strip", "")
        veq, unit = _vol_equiv(cc, qty, strip0, tt)

        d = cc_data[cc]
        d["trades"] += 1
        d["unit"]    = unit
        if tt in _OUTRIGHT_TYPES:
            d["out_vol"] += qty
            d["out_veq"] += veq
        elif tt in _SPREAD_TYPES:
            d["spr_vol"] += qty
            d["spr_veq"] += veq

    # Clear sheet (row 3+)
    for rw in ws2.iter_rows(min_row=3):
        for c in rw:
            c.value  = None
            c.fill   = F_OUT
            c.border = Border()
            c.font   = FN_BODY

    _hdr(ws2, 2, _SUMMARY_HDR_COLS)
    for col, w in _SUMMARY_HDR_WIDTHS.items():
        ws2.column_dimensions[col].width = w

    r  = 3
    nc = len(_SUMMARY_HDR_COLS)

    def _sum_cell(row_num, vals, fill, bold=False, height=16, white_text=False):
        aligns = ["left"] + ["right"] * (nc - 1)
        font = FN_HDR if white_text else None
        for ci, (val, ha) in enumerate(zip(vals, aligns), 1):
            _c(ws2, row_num, ci, val, fill=fill, halign=ha, bold=bold, font=font)
        ws2.row_dimensions[row_num].height = height

    def _banner(row_num, text, height=20):
        for ci in range(1, nc + 1):
            c = ws2.cell(row=row_num, column=ci,
                         value=(text if ci == 1 else None))
            c.fill      = F_CC_BAND
            c.font      = FN_HDR
            c.border    = BORD
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[row_num].height = height

    grand = _blank_cc()

    for cat_name, cat_ccs in _SUMMARY_CATEGORIES:
        # Determine which CCs belong to this category
        if cat_ccs is None:
            ccs_in_cat = sorted(cc for cc in cc_data if cc not in _ALL_FREIGHT_CC)
        else:
            ccs_in_cat = sorted(cc for cc in cat_ccs if cc in cc_data)

        if not ccs_in_cat:
            continue

        _banner(r, cat_name)
        r += 1

        cat = _blank_cc()
        units_seen: set[str] = set()

        for cc in ccs_in_cat:
            d = cc_data[cc]
            tot_vol = d["out_vol"] + d["spr_vol"]
            tot_veq = d["out_veq"] + d["spr_veq"]
            unit    = d["unit"]
            units_seen.add(unit)
            _sum_cell(r, [cc, unit,
                          d["out_vol"] or None, d["out_veq"] or None,
                          d["spr_vol"] or None, d["spr_veq"] or None,
                          tot_vol, tot_veq, d["trades"]],
                      F_OUT, height=16)
            r += 1
            for k in ("out_vol", "out_veq", "spr_vol", "spr_veq", "trades"):
                cat[k] += d[k]

        cat_tot_vol = cat["out_vol"] + cat["spr_vol"]
        cat_tot_veq = cat["out_veq"] + cat["spr_veq"]
        cat_unit = next(iter(units_seen)) if len(units_seen) == 1 else "mixed"
        _sum_cell(r, [f"  Total {cat_name}", cat_unit,
                      cat["out_vol"] or None, cat["out_veq"] or None,
                      cat["spr_vol"] or None, cat["spr_veq"] or None,
                      cat_tot_vol, cat_tot_veq, int(cat["trades"])],
                  F_SUMMARY, bold=True, height=18)
        r += 2

        for k in ("out_vol", "out_veq", "spr_vol", "spr_veq", "trades"):
            grand[k] += cat[k]

    # Grand Total — vol equiv omitted (mixed units across categories)
    gt_vol = grand["out_vol"] + grand["spr_vol"]
    _sum_cell(r, ["  GRAND TOTAL", "KT",
                  grand["out_vol"] or None, "—",
                  grand["spr_vol"] or None, "—",
                  gt_vol, "—", int(grand["trades"])],
              F_GRAND, bold=True, height=22, white_text=True)


# ═══════════════════════════════════════════════════════════════════════════
#  APPEND TRADES + SAVE
# ═══════════════════════════════════════════════════════════════════════════

def append_trades(path: str, trades: list[dict],
                  raw_row_count: int, source_file: str):
    existing = load_existing_keys(path)
    wb = load_workbook(path)
    ws = wb[SH_LOG]
    wi = wb[SH_IMPORT]

    # Ensure Day Summary sheet exists (for workbooks created before this feature)
    if SH_SUMMARY not in wb.sheetnames:
        ws2 = wb.create_sheet(SH_SUMMARY)
        ws2.freeze_panes = "A3"
        ws2["A1"] = "Day Summary  —  Freight volumes by category"
        ws2["A1"].font      = FN_TITLE
        ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[1].height = 28

    new_count  = 0
    skip_count = 0

    for t in trades:
        legs  = t.get("legs", [])
        l1    = legs[0]["strip"] if legs else ""
        key   = (t["timestamp"].strip(), t["trade_type"],
                 t["cc"].strip(), l1)
        if key in existing:
            skip_count += 1
            continue
        _write_log_row(ws, t, source_file)
        existing.add(key)
        new_count += 1

    _rebuild_tally(wb)
    _build_day_summary(wb)

    log_r = wi.max_row + 1
    for ci, val in enumerate([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        source_file, raw_row_count, len(trades), new_count, skip_count
    ], 1):
        _c(wi, log_r, ci, val,
           halign="left" if ci <= 2 else "center")
    wi.row_dimensions[log_r].height = 17

    wb.save(path)
    return new_count, skip_count


# ═══════════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Trade blotter screengrab → Excel accumulator")
    ap.add_argument("image",
                    help="Path to screengrab PNG/JPG")
    ap.add_argument("--output", "-o", default="trade_tally.xlsx",
                    help="Excel file (created on first run, updated thereafter)")
    ap.add_argument("--api-key",
                    default=os.environ.get("ANTHROPIC_API_KEY", ""),
                    help="Anthropic API key (or set ANTHROPIC_API_KEY env var)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Parse and print trades without writing to Excel")
    args = ap.parse_args()

    if not args.api_key:
        sys.exit("ERROR: No API key found. "
                 "Set ANTHROPIC_API_KEY or pass --api-key.")
    if not Path(args.image).exists():
        sys.exit(f"ERROR: Image not found: {args.image}")

    print(f"📷  Parsing {args.image} …")
    raw_rows = parse_image_with_claude(args.image, args.api_key)
    print(f"✅  {len(raw_rows)} raw row(s) extracted.")

    trades = group_rows_into_trades(raw_rows)
    print(f"🔗  Grouped into {len(trades)} trade(s).")

    if args.dry_run:
        print(json.dumps(trades, indent=2))
        return

    if not Path(args.output).exists():
        print(f"📄  Creating new workbook: {args.output}")
        build_fresh_workbook(args.output)

    added, skipped = append_trades(
        args.output, trades, len(raw_rows), Path(args.image).name)

    print(f"📊  {added} new trade(s) added, {skipped} duplicate(s) skipped.")
    print(f"💾  Saved → {args.output}")


if __name__ == "__main__":
    main()
